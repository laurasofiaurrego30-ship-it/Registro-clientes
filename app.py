from flask import Flask, request, render_template_string, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Nombre del archivo Excel donde se guardarán los registros
EXCEL_FILE = "clientes.xlsx"

# Si no existe, lo crea con encabezados
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Cédula", "Nombres", "Apellidos", "Dirección", "Teléfono", "Correo"])
    wb.save(EXCEL_FILE)

# Ruta principal que carga tu formulario
@app.route("/")
def index():
    with open("registro.html", encoding="utf-8") as f:
        return f.read()

# Ruta para procesar el registro
@app.route("/procesar-registro", methods=["POST"])
def procesar_registro():
    cedula = request.form["cedula"]
    nombres = request.form["nombres"]
    apellidos = request.form["apellidos"]
    direccion = request.form["direccion"]
    telefono = request.form["telefono"]
    correo = request.form["correo"]

    # Abrir el Excel y guardar los datos
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([cedula, nombres, apellidos, direccion, telefono, correo])
    wb.save(EXCEL_FILE)

    # Mensaje de confirmación con botón para regresar
    return render_template_string("""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Registro Exitoso</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background-color: #f3f3f3;
                text-align: center;
                padding: 50px;
            }
            .card {
                background: white;
                padding: 30px;
                border-radius: 10px;
                display: inline-block;
                box-shadow: 0px 4px 12px rgba(0,0,0,0.1);
            }
            h2 {
                color: #3eb489;
                margin-bottom: 20px;
            }
            a {
                display: inline-block;
                background: #3eb489;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                text-decoration: none;
                font-size: 16px;
            }
            a:hover {
                background: #34a67a;
            }
        </style>
    </head>
    <body>
        <div class="card">
            <h2>✅ Cliente registrado con éxito</h2>
            <a href="/">Volver al formulario</a>
        </div>
    </body>
    </html>
    """)
    
if __name__ == "__main__":
    app.run(debug=True)
